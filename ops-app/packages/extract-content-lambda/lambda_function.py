import logging
import tempfile
import os
from io import StringIO
from docx import Document
from openpyxl import load_workbook
import csv
import boto3
import json
from sentry_sdk.integrations.aws_lambda import AwsLambdaIntegration
import sentry_sdk

logger = logging.getLogger()
logger.setLevel(logging.INFO)

sentry_sdk.init(
    dsn="https://00651930388f84074cf22192f8410569@o4505774316060672.ingest.sentry.io/4505802562338816",
    integrations=[AwsLambdaIntegration()],
    # Set traces_sample_rate to 1.0 to capture 100%
    # of transactions for performance monitoring.
    # We recommend adjusting this value in production.
    traces_sample_rate=0,
    # Set profiles_sample_rate to 1.0 to profile 100%
    # of sampled transactions.
    # We recommend adjusting this value in production.
    profiles_sample_rate=0,
)

s3 = boto3.client('s3')

PAGE_DELIMITER = '-' * 30 + '@@' + '-' * 30


def handler(event, context):
    # Get bucket and object key from the Lambda event trigger
    bucket = event['Records'][0]['s3']['bucket']['name']
    key = event['Records'][0]['s3']['object']['key']

    logger.info(f"Converting {key} in bucket {bucket}")

    # Determine the file type from the file extension
    file_type = key.split('.')[-1]
    # file_name =  os.path.basename(key)

    with tempfile.TemporaryDirectory() as tmp_dir:
        download_path = os.path.join(tmp_dir, f"downloaded_file.{file_type}")
        upload_path = os.path.join(tmp_dir, 'extracted_file')

        s3.download_file(bucket, key, download_path)

        if file_type in ['doc', 'docx']:
            convert_docx_to_text(download_path, upload_path)
        elif file_type == 'xlsx':
            convert_xlsx_to_csv(download_path, upload_path)

        new_key = f"{key}.extracted"
        s3.upload_file(upload_path, bucket, new_key)


def convert_docx_to_text(read_path, extracted_path):
    doc = Document(read_path)
    with open(extracted_path, 'w') as extracted_file:
        for para in doc.paragraphs:
            extracted_file.write(para.text)
            extracted_file.write('\n\n')


def convert_xlsx_to_csv(read_path, extracted_path):
    wb = load_workbook(read_path)
    with open(extracted_path, 'w') as extracted_file:
        for sheet in wb.worksheets:
            extracted_file.write(PAGE_DELIMITER + '\n')
            meta_info = {
                'sheetTitle': sheet.title
            }
            extracted_file.write(f'META:{json.dumps(meta_info)}\n')

            output = StringIO()
            writer = csv.writer(output)
            for row in sheet.iter_rows():
                writer.writerow([cell.value for cell in row])

            extracted_file.write(output.getvalue())
